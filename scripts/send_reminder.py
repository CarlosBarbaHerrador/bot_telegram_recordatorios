import urllib.request
import urllib.parse
import openpyxl
import os
import sys
import math
import tempfile
import time
from datetime import date

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

EXCEL_URL = "https://docs.google.com/spreadsheets/d/1fL7CZ_Td2JAb0Q5RlL_plOMLqrZGl6Ax1zbXJa_kGaE/export?format=xlsx"
SHEET_NAME = " Nigun Grid Luin"

SECTIONS = {
    'Guardia de No Muertos': 'guardia',
    'Apostatas de Myrkull': 'apostatas',
    'Basura': 'basura',
    'Mis Levantados': 'levantados',
}

SKIP_NAMES = {'Total'}

def download_excel(filepath, retries=3, delay=5):
    for attempt in range(retries):
        try:
            urllib.request.urlretrieve(EXCEL_URL, filepath)
            return
        except Exception as e:
            if attempt < retries - 1:
                wait = delay * (2 ** attempt)
                print(f"Error descargando (intento {attempt+1}/{retries}): {e}. Reintentando en {wait}s...", file=sys.stderr)
                time.sleep(wait)
            else:
                raise

def is_valid_number(v):
    return isinstance(v, (int, float)) and not math.isnan(v)

def parse_sections(ws):
    section_data = {'guardia': [], 'apostatas': [], 'basura': [], 'levantados': []}
    current_section = None

    for row in ws.iter_rows(min_row=1, values_only=True):
        row_list = list(row)
        c23 = row_list[22] if len(row_list) > 22 else None
        c23_str = str(c23).strip() if c23 is not None else None

        if c23_str in SECTIONS:
            current_section = SECTIONS[c23_str]
            continue

        if current_section is None:
            continue

        if is_valid_number(c23):
            qty = int(c23)
            name = str(row_list[23]).strip() if len(row_list) > 23 and row_list[23] is not None else ''
            if name not in SKIP_NAMES and name:
                section_data[current_section].append((qty, name))

    merged = {}
    for section in section_data.values():
        for qty, name in section:
            merged[name] = merged.get(name, 0) + qty

    return merged

CURRENCY_KEYWORDS = {'Oro', 'Plata', 'Bronce'}

def parse_currency(ws):
    currencies = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        row_list = list(row)
        for i, cell in enumerate(row_list):
            if cell is not None and str(cell).strip() in CURRENCY_KEYWORDS:
                ctype = str(cell).strip()
                amount = None
                if i > 0 and is_valid_number(row_list[i-1]):
                    amount = int(row_list[i-1])
                elif i + 1 < len(row_list) and is_valid_number(row_list[i+1]):
                    amount = int(row_list[i+1])
                if amount is not None:
                    currencies[ctype] = amount
    return currencies

def read_ganancia():
    path = os.path.join(os.path.dirname(__file__), "..", "ganancia.txt")
    with open(path, encoding="utf-8") as f:
        return f.read().strip()

def build_message(merged, currencies, ganancia):
    gran_total = sum(merged.values())
    sorted_items = sorted(merged.items(), key=lambda x: x[1], reverse=True)

    sep = "-" * 30

    today = date.today().strftime("%d/%m/%Y (%A)")
    dias = {"Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miércoles",
            "Thursday": "Jueves", "Friday": "Viernes", "Saturday": "Sábado", "Sunday": "Domingo"}
    for en, es in dias.items():
        today = today.replace(en, es)

    ganancia_lines = ganancia.split('\n')
    undead_gains = []
    dinero_gains = {}
    in_money = False
    for l in ganancia_lines:
        ls = l.strip()
        if ls == "--- Dinero ---":
            in_money = True
            continue
        if in_money:
            for kw in CURRENCY_KEYWORDS:
                if kw in ls:
                    parts = ls.split()
                    for part in parts:
                        try:
                            dinero_gains[kw] = dinero_gains.get(kw, 0) + int(part)
                        except ValueError:
                            pass
                    break
        else:
            if ls:
                undead_gains.append(ls)

    lines = []
    lines.append(f"Periodo de Ningun - {today}")
    lines.append("")

    lines.append("---- No Muertos ----")
    for name, qty in sorted_items:
        lines.append(f"{qty} {name}")
    lines.append(sep)
    lines.append(f"Total: {gran_total}")
    lines.append("")

    if currencies:
        lines.append("---- Dinero ----")
        for ctype in ('Oro', 'Plata', 'Bronce'):
            if ctype in currencies:
                lines.append(f"{currencies[ctype]} de {ctype}")
    lines.append("=" * 40)

    lines.append("---- No Muertos ----")
    for l in undead_gains:
        lines.append(l)
    lines.append("")

    if dinero_gains:
        lines.append("---- Dinero ----")
        for ctype in ('Oro', 'Plata', 'Bronce'):
            if ctype in dinero_gains:
                lines.append(f"+ {dinero_gains[ctype]} {ctype}")

    if currencies:
        lines.append("---- Total Dinero ----")
        combined = {}
        for ctype in currencies:
            combined[ctype] = currencies[ctype]
        for ctype, amount in dinero_gains.items():
            combined[ctype] = combined.get(ctype, 0) + amount
        for ctype in ('Oro', 'Plata', 'Bronce'):
            if ctype in combined:
                lines.append(f"{combined[ctype]} de {ctype}")

    return "\n".join(lines)

def send_telegram(token, chat_id, message):
    data = urllib.parse.urlencode({'chat_id': chat_id, 'text': message, 'parse_mode': 'Markdown'}).encode()
    req = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data=data,
        headers={'Content-Type': 'application/x-www-form-urlencoded'}
    )
    with urllib.request.urlopen(req) as resp:
        return resp.status

ERROR_LOCK = os.path.join(tempfile.gettempdir(), "periodo_error.lock")
ERROR_COOLDOWN = 3600

def can_send_error():
    try:
        if os.path.exists(ERROR_LOCK):
            with open(ERROR_LOCK) as f:
                last = float(f.read().strip())
            if time.time() - last < ERROR_COOLDOWN:
                return False
        return True
    except Exception:
        return True

def mark_error_sent():
    try:
        with open(ERROR_LOCK, "w") as f:
            f.write(str(time.time()))
    except Exception:
        pass

def send_error_alert(token, chat_id, error_msg):
    if not can_send_error():
        print("Error alert omitido (cooldown de 1h)", file=sys.stderr)
        return
    text = f"*⚠️ Error en Periodo Bot:*\n{error_msg}"
    try:
        data = urllib.parse.urlencode({'chat_id': chat_id, 'text': text, 'parse_mode': 'Markdown'}).encode()
        req = urllib.request.Request(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data=data,
            headers={'Content-Type': 'application/x-www-form-urlencoded'}
        )
        with urllib.request.urlopen(req, timeout=15):
            pass
        mark_error_sent()
    except Exception:
        pass

def load_dotenv():
    path = os.path.join(os.path.dirname(__file__), "..", ".env")
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    key, _, val = line.partition("=")
                    if key not in os.environ:
                        os.environ[key.strip()] = val.strip()

def should_run_today():
    now = time.gmtime()
    weekday = now.tm_wday
    hour = now.tm_hour
    if weekday not in (6, 1, 4):
        return False
    if hour < 13:
        return False
    return True

def main():
    if not should_run_today():
        return
    load_dotenv()
    token = os.environ.get('TELEGRAM_TOKEN')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID')
    if not token or not chat_id:
        print("Faltan TELEGRAM_TOKEN o TELEGRAM_CHAT_ID", file=sys.stderr)
        sys.exit(1)

    filepath = os.path.join(tempfile.gettempdir(), "ningun.xlsx")
    try:
        download_excel(filepath)
    except Exception as e:
        msg = f"No pude descargar el Excel: {e}"
        print(msg, file=sys.stderr)
        send_error_alert(token, chat_id, msg)
        sys.exit(1)

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[SHEET_NAME]

        merged = parse_sections(ws)
        currencies = parse_currency(ws)
        wb.close()
    except Exception as e:
        msg = f"No pude leer el Excel (cambio de estructura?): {e}"
        print(msg, file=sys.stderr)
        send_error_alert(token, chat_id, msg)
        sys.exit(1)

    if not merged:
        msg = "No encontré ninguna sección de no-muertos en el Excel. ¿Cambiaste la estructura?"
        print(msg, file=sys.stderr)
        send_error_alert(token, chat_id, msg)
        sys.exit(1)

    message = build_message(merged, currencies, read_ganancia())
    print("Mensaje generado:")
    print(message)

    status = send_telegram(token, chat_id, message)
    print(f"Telegram response: {status}")

if __name__ == '__main__':
    main()
