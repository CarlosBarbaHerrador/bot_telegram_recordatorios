import urllib.request
import urllib.parse
import json
import os
import sys
import time
import math
import tempfile
from datetime import date

sys.stdout.reconfigure(encoding='utf-8', errors='replace')
sys.stderr.reconfigure(encoding='utf-8', errors='replace')

EXCEL_URL = "https://docs.google.com/spreadsheets/d/1fL7CZ_Td2JAb0Q5RlL_plOMLqrZGl6Ax1zbXJa_kGaE/export?format=xlsx"
SHEET_NAME = " Nigun Grid Luin"

SECTIONS = {
    'Guardia de No Muertos': 'guardia',
    'Mis Levantados en el Libro': 'levantados',
    'No muertos de Radija': 'radija',
    'Muertos con ordenes permanentes': 'permanentes',
}

SKIP_NAMES = {'Total'}
CURRENCY_KEYWORDS = {'Oro', 'Plata', 'Bronce'}


def fetch_json(url):
    with urllib.request.urlopen(url) as r:
        return json.loads(r.read().decode())


def post_json(url, data):
    body = json.dumps(data).encode()
    req = urllib.request.Request(url, data=body, headers={'Content-Type': 'application/json'})
    with urllib.request.urlopen(req) as r:
        return json.loads(r.read().decode())


def download_excel(filepath):
    urllib.request.urlretrieve(EXCEL_URL, filepath)


def is_valid_number(v):
    return isinstance(v, (int, float)) and not math.isnan(v)


def parse_sections(ws):
    import openpyxl
    merged = {}
    current_section = None

    col_pairs = [(22, 23), (28, 29)]

    for row in ws.iter_rows(min_row=1, values_only=True):
        row_list = list(row)
        c22 = str(row_list[22]).strip() if len(row_list) > 22 and row_list[22] is not None else ''
        c28 = str(row_list[28]).strip() if len(row_list) > 28 and row_list[28] is not None else ''

        found_section = None
        if c22 in SECTIONS:
            found_section = SECTIONS[c22]
        elif c28 in SECTIONS:
            found_section = SECTIONS[c28]

        if found_section is not None:
            current_section = found_section
            continue

        if current_section is None:
            continue

        for qty_col, name_col in col_pairs:
            if len(row_list) <= qty_col or len(row_list) <= name_col:
                continue
            qty_val = row_list[qty_col]
            name_val = row_list[name_col]
            if is_valid_number(qty_val):
                qty = int(qty_val)
                name = str(name_val).strip() if name_val is not None else ''
                if name not in SKIP_NAMES and name:
                    merged[name] = merged.get(name, 0) + qty

    return merged


def parse_currency(ws):
    currencies = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        row_list = list(row)
        for i, cell in enumerate(row_list):
            if cell is not None and str(cell).strip() in CURRENCY_KEYWORDS:
                ctype = str(cell).strip()
                amount = None
                if i > 0 and is_valid_number(row_list[i-1]):
                    amount = int(row_list[i-1])
                elif i + 1 < len(row_list) and is_valid_number(row_list[i+1]):
                    amount = int(row_list[i+1])
                if amount is not None:
                    currencies[ctype] = amount
    return currencies


def read_ganancia():
    path = os.path.join(os.path.dirname(__file__), "..", "ganancia.txt")
    with open(path, encoding="utf-8") as f:
        return f.read().strip()


def build_message(merged, currencies, ganancia):
    gran_total = sum(merged.values())
    sorted_items = sorted(merged.items(), key=lambda x: x[1], reverse=True)
    sep = "-" * 30
    today = date.today().strftime("%d/%m/%Y (%A)")
    dias = {"Monday": "Lunes", "Tuesday": "Martes", "Wednesday": "Miércoles",
            "Thursday": "Jueves", "Friday": "Viernes", "Saturday": "Sábado", "Sunday": "Domingo"}
    for en, es in dias.items():
        today = today.replace(en, es)

    ganancia_lines = ganancia.split('\n')
    undead_gains = []
    dinero_gains = {}
    in_money = False
    for l in ganancia_lines:
        ls = l.strip()
        if ls == "--- Dinero ---":
            in_money = True
            continue
        if in_money:
            for kw in CURRENCY_KEYWORDS:
                if kw in ls:
                    parts = ls.split()
                    for part in parts:
                        try:
                            dinero_gains[kw] = dinero_gains.get(kw, 0) + int(part)
                        except ValueError:
                            pass
                    break
        else:
            if ls:
                undead_gains.append(ls)

    lines = []
    lines.append(f"Periodo de Ningun - {today}")
    lines.append("")
    lines.append("---- No Muertos ----")
    for name, qty in sorted_items:
        lines.append(f"{qty} {name}")
    lines.append(sep)
    lines.append(f"Total: {gran_total}")
    lines.append("")
    if currencies:
        lines.append("---- Dinero ----")
        for ctype in ('Oro', 'Plata', 'Bronce'):
            if ctype in currencies:
                lines.append(f"{currencies[ctype]} de {ctype}")
    lines.append("=" * 40)
    lines.append("---- No Muertos ----")
    for l in undead_gains:
        lines.append(l)
    lines.append("")
    if dinero_gains:
        lines.append("---- Dinero ----")
        for ctype in ('Oro', 'Plata', 'Bronce'):
            if ctype in dinero_gains:
                lines.append(f"+ {dinero_gains[ctype]} {ctype}")
    if currencies:
        lines.append("---- Total Dinero ----")
        combined = {}
        for ctype in currencies:
            combined[ctype] = currencies[ctype]
        for ctype, amount in dinero_gains.items():
            combined[ctype] = combined.get(ctype, 0) + amount
        for ctype in ('Oro', 'Plata', 'Bronce'):
            if ctype in combined:
                lines.append(f"{combined[ctype]} de {ctype}")
    return "\n".join(lines)


def send_message(token, chat_id, text):
    data = urllib.parse.urlencode({'chat_id': chat_id, 'text': text, 'parse_mode': 'Markdown'}).encode()
    req = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data=data,
        headers={'Content-Type': 'application/x-www-form-urlencoded'}
    )
    with urllib.request.urlopen(req) as resp:
        return resp.status


def load_dotenv():
    path = os.path.join(os.path.dirname(__file__), "..", ".env")
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    key, _, val = line.partition("=")
                    if key not in os.environ:
                        os.environ[key.strip()] = val.strip()


def generate_and_send(token, chat_id):
    filepath = os.path.join(tempfile.gettempdir(), "ningun_periodo.xlsx")
    try:
        download_excel(filepath)
    except Exception as e:
        send_message(token, chat_id, f"*⚠️ Error:* No pude descargar el Excel: {e}")
        return
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        ws = wb[SHEET_NAME]
        merged = parse_sections(ws)
        currencies = parse_currency(ws)
        wb.close()
    except Exception as e:
        send_message(token, chat_id, f"*⚠️ Error:* No pude leer el Excel: {e}")
        return
    if not merged:
        send_message(token, chat_id, "*⚠️ Error:* No encontré secciones de no-muertos en el Excel.")
        return
    message = build_message(merged, currencies, read_ganancia())
    status = send_message(token, chat_id, message)
    print(f"Periodo enviado, response: {status}")


def main():
    load_dotenv()
    token = os.environ.get('TELEGRAM_TOKEN')
    chat_id = os.environ.get('TELEGRAM_CHAT_ID')
    if not token or not chat_id:
        print("Faltan TELEGRAM_TOKEN o TELEGRAM_CHAT_ID", file=sys.stderr)
        sys.exit(1)

    offset = 0
    print("🤖 Periodo Bot iniciado. Escuchando /periodo o !periodo...")
    while True:
        try:
            url = f"https://api.telegram.org/bot{token}/getUpdates?offset={offset}&timeout=60"
            updates = fetch_json(url)
            for upd in updates.get('result', []):
                offset = upd['update_id'] + 1
                msg = upd.get('message')
                if not msg:
                    continue
                text = (msg.get('text') or '').strip().lower()
                if text in ('/periodo', '!periodo'):
                    print(f"Comando recibido de {msg['from'].get('username', '?')}")
                    send_message(token, msg['chat']['id'], "⏳ Generando periodo...")
                    generate_and_send(token, msg['chat']['id'])
                    print("✅ Periodo enviado.")
        except KeyboardInterrupt:
            print("Bot detenido.")
            break
        except Exception as e:
            print(f"Error en polling: {e}", file=sys.stderr)
            time.sleep(5)


if __name__ == '__main__':
    main()
